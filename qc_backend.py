import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import re
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile


def clean_filename(text):
    return re.sub(r'[\\/*?:"<>|]', "", str(text))

def build_zip(file_list):
    zip_buffer = BytesIO()

    with ZipFile(zip_buffer, "w") as zip_file:
        for filename, file_buffer in file_list:
            zip_file.writestr(filename, file_buffer.getvalue())

    zip_buffer.seek(0)
    return zip_buffer


def run_qc(uploaded_file, assignee, qc_type):

    df = pd.read_excel(uploaded_file)
    df.columns = df.columns.str.strip()

    template_path = Path("templates") / "Doc QC Template.xlsx"
    if not template_path.exists():
        raise FileNotFoundError("Template not found in /templates folder")
    
    assignee_map = {
        "denise": ("Denise Bahena", "DB"),
        "adam": ("Adam Bell", "AB"),
        "john": ("John Endslow", "JE"),
        "sandra": ("Sandra Sin", "SS"),
    }

    fullname, initials = assignee_map.get(assignee.lower(), (assignee, "XX"))
    today = datetime.today().strftime("%m-%d-%Y")

    files = []

    if qc_type == "Mods / Renewals / Extensions":
        df = df[df["Assigned To"].fillna("").str.lower().str.contains(assignee.lower())]

    elif qc_type == "New Loan":
        df = df[df["Doc QC Assigned To"].fillna("").str.lower().str.contains(assignee.lower())]

    else:
        raise ValueError("Invalid QC type")

    for _, row in df.iterrows():

        wb = load_workbook(template_path)
        ws = wb["Archive Checklist"]

        ws["C3"] = row.get("DealNumber", "")
        ws["C4"] = row.get("AccountNumber", "")
        ws["C5"] = row.get("FullName", "")
        ws["C7"] = fullname
        ws["C8"] = today
        ws["C163"] = fullname

        if qc_type == "Mods / Renewals / Extensions":
            ws["G1"] = row.get("Stage Date", "")
        else:
            ws["G1"] = row.get("Stage Date (UTC)", "")

        filename = (
            f"{clean_filename(row.get('FullName','Unknown'))}_BH_"
            f"{clean_filename(row.get('AccountNumber','Unknown'))}_"
            f"{today}_{initials}.xlsx"
        )

        # ============================
        # SAVE TO MEMORY (NOT DISK)
        # ============================
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        files.append((filename, buffer))

    return files